import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("Transactions.xlsx")
sheet = wb["Sheet1"]

cell1 = sheet["a1"]

print(cell1.value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    discounted_price  = cell.value * 0.9
    discounted_price_cell = sheet.cell(row, 4)
    discounted_price.v
    print(cell.value)

wb.save("Transaction2.xlsx")

